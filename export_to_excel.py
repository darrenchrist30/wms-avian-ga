"""
Export GA Results to Excel
Membuat dokumentasi lengkap hasil GA dalam format Excel
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import random
import numpy as np

# Run the GA comparison
import sys
sys.path.append('.')

print("Generating comprehensive Excel documentation...")
print("="*80)

# ============================================================================
# DATA PREPARATION
# ============================================================================

# Results from previous runs
results_25_sku = {
    'configs': [
        'roulette+uniform+swap',
        'tournament+one_point+swap',
        'tournament+two_point+swap',
        'tournament+uniform+swap',
        'roulette+one_point+swap',
        'roulette+two_point+swap',
        'tournament+two_point+random',
        'tournament+uniform+random',
        'tournament+one_point+random',
        'roulette+one_point+random',
        'roulette+two_point+random',
        'roulette+uniform+random',
    ],
    'fitness': [2340, 2340, 2340, 2340, 2310, 2310, 2300, 2270, 2265, 2160, 2135, 2065],
    'max_possible': 2500,
    'num_skus': 25
}

results_100_sku = {
    'roulette': {
        'fitness': 8315,
        'achievement': 83.2,
        'time': 45.2,
        'convergence_gen': 180
    },
    'tournament': {
        'fitness': 8470,
        'achievement': 84.7,
        'time': 44.8,
        'convergence_gen': 20
    },
    'max_possible': 10000,
    'num_skus': 100
}

# ============================================================================
# CREATE EXCEL WORKBOOK
# ============================================================================

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="0D8564", end_color="0D8564", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=11)
highlight_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# SHEET 1: RINGKASAN EKSEKUTIF
# ============================================================================

ws1 = wb.create_sheet("1. Executive Summary")

# Title
ws1['A1'] = "GENETIC ALGORITHM - WAREHOUSE STORAGE OPTIMIZATION"
ws1['A1'].font = Font(bold=True, size=16, color="004230")
ws1.merge_cells('A1:F1')

ws1['A2'] = "Perbandingan Operator Seleksi, Crossover, dan Mutasi"
ws1['A2'].font = Font(size=12, italic=True)
ws1.merge_cells('A2:F2')

# Recommendation
ws1['A4'] = "🏆 REKOMENDASI KONFIGURASI OPTIMAL"
ws1['A4'].font = Font(bold=True, size=14, color="0D8564")
ws1.merge_cells('A4:F4')

ws1['A6'] = "Untuk WMS dengan volume tinggi (50+ SKU) dan constraint ketat:"
ws1.merge_cells('A6:F6')

data = [
    ['Komponen', 'Metode Terpilih', 'Alasan'],
    ['Selection', 'TOURNAMENT', 'Convergence 10x lebih cepat, fitness 155 poin lebih tinggi'],
    ['Crossover', 'UNIFORM', 'Maximum exploration, cocok untuk constraint-heavy problems'],
    ['Mutation', 'SWAP', 'Maintain feasibility, 275 poin lebih baik dari random'],
    ['Tournament Size', '3', 'Balance antara selection pressure dan diversity'],
    ['Population Size', '100-150', 'Cukup untuk diverse solution space'],
    ['Generations', '150-200', 'Ensure full convergence'],
    ['Mutation Rate', '0.15', 'Optimal untuk exploration tanpa disruptive'],
]

for i, row in enumerate(data, start=8):
    for j, value in enumerate(row, start=1):
        cell = ws1.cell(row=i, column=j, value=value)
        cell.border = border
        if i == 8:  # Header row
            cell.fill = subheader_fill
            cell.font = subheader_font
        if j == 2 and i > 8:  # Metode terpilih
            cell.font = Font(bold=True, color="0D8564")

# Performance Summary
ws1['A17'] = "📊 PERFORMANCE SUMMARY"
ws1['A17'].font = Font(bold=True, size=14, color="0D8564")
ws1.merge_cells('A17:F17')

summary_data = [
    ['Skenario', 'Konfigurasi', 'Fitness', 'Max Possible', 'Achievement', 'Convergence'],
    ['25 SKUs (Challenging)', 'Tournament+Uniform+Swap', '2340', '2500', '93.6%', 'Gen 60'],
    ['100 SKUs (Extreme)', 'Tournament+Uniform+Swap', '8470', '10000', '84.7%', 'Gen 20'],
]

for i, row in enumerate(summary_data, start=19):
    for j, value in enumerate(row, start=1):
        cell = ws1.cell(row=i, column=j, value=value)
        cell.border = border
        if i == 19:
            cell.fill = subheader_fill
            cell.font = subheader_font
        if 'Tournament' in str(value):
            cell.fill = highlight_fill
            cell.font = Font(bold=True)

# Adjust column widths
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 20

# ============================================================================
# SHEET 2: HASIL 25 SKU
# ============================================================================

ws2 = wb.create_sheet("2. Results 25 SKUs")

ws2['A1'] = "HASIL EKSPERIMEN: 25 SKUs Incoming"
ws2['A1'].font = Font(bold=True, size=14, color="004230")
ws2.merge_cells('A1:E1')

ws2['A3'] = "Skenario:"
ws2['B3'] = "25 SKU incoming, 21 cells, usage 76.8%, space limited"
ws2.merge_cells('B3:E3')

# Results table
headers = ['Rank', 'Konfigurasi', 'Fitness', 'Achievement (%)', 'Status']
for j, header in enumerate(headers, start=1):
    cell = ws2.cell(row=5, column=j, value=header)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.border = border

for i, (config, fitness) in enumerate(zip(results_25_sku['configs'], results_25_sku['fitness']), start=6):
    achievement = (fitness / results_25_sku['max_possible']) * 100
    status = '🏆 TOP' if i <= 9 else ''

    row_data = [i-5, config, fitness, f"{achievement:.1f}%", status]
    for j, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=i, column=j, value=value)
        cell.border = border
        if i <= 9 and achievement >= 93.6:  # Top performers
            cell.fill = PatternFill(start_color="E7F4E4", end_color="E7F4E4", fill_type="solid")

# Analysis
ws2['A19'] = "📈 ANALISIS"
ws2['A19'].font = Font(bold=True, size=12, color="0D8564")

analysis_points = [
    "• 4 konfigurasi mencapai fitness maksimal 2340 (93.6%)",
    "• Swap mutation WAJIB: 275 poin lebih baik dari random",
    "• Uniform crossover memberikan exploration terbaik",
    "• Tournament & Roulette setara untuk kasus 25 SKU",
]

for i, point in enumerate(analysis_points, start=20):
    ws2.cell(row=i, column=1, value=point)
    ws2.merge_cells(f'A{i}:E{i}')

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 15

# ============================================================================
# SHEET 3: HASIL 100 SKU
# ============================================================================

ws3 = wb.create_sheet("3. Results 100 SKUs")

ws3['A1'] = "HASIL EKSPERIMEN: 100 SKUs Incoming (EXTREME)"
ws3['A1'].font = Font(bold=True, size=14, color="004230")
ws3.merge_cells('A1:F1')

ws3['A3'] = "Skenario:"
ws3['B3'] = "100 SKU incoming, 50 cells, usage 85%, TIGHT constraint"
ws3.merge_cells('B3:F3')

# Head-to-head comparison
comparison_headers = ['Metode', 'Fitness', 'Achievement', 'Convergence', 'Time (s)', 'Winner']
for j, header in enumerate(comparison_headers, start=1):
    cell = ws3.cell(row=5, column=j, value=header)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.border = border

# Roulette row
roulette_data = [
    'Roulette+Uniform+Swap',
    results_100_sku['roulette']['fitness'],
    f"{results_100_sku['roulette']['achievement']:.1f}%",
    f"Gen {results_100_sku['roulette']['convergence_gen']}",
    results_100_sku['roulette']['time'],
    ''
]
for j, value in enumerate(roulette_data, start=1):
    cell = ws3.cell(row=6, column=j, value=value)
    cell.border = border

# Tournament row
tournament_data = [
    'Tournament+Uniform+Swap',
    results_100_sku['tournament']['fitness'],
    f"{results_100_sku['tournament']['achievement']:.1f}%",
    f"Gen {results_100_sku['tournament']['convergence_gen']}",
    results_100_sku['tournament']['time'],
    '🏆 WINNER'
]
for j, value in enumerate(tournament_data, start=1):
    cell = ws3.cell(row=7, column=j, value=value)
    cell.border = border
    cell.fill = highlight_fill
    if j == 6:
        cell.font = Font(bold=True, size=12)

# Difference
diff_fitness = results_100_sku['tournament']['fitness'] - results_100_sku['roulette']['fitness']
diff_gen = results_100_sku['roulette']['convergence_gen'] - results_100_sku['tournament']['convergence_gen']

ws3['A9'] = "Selisih:"
ws3['B9'] = f"+{diff_fitness} poin (1.5%)"
ws3['C9'] = "Convergence:"
ws3['D9'] = f"{diff_gen}x lebih cepat"
ws3['D9'].font = Font(bold=True, color="0D8564")

# Key Insights
ws3['A12'] = "🔍 KEY INSIGHTS"
ws3['A12'].font = Font(bold=True, size=12, color="0D8564")

insights = [
    "Tournament Selection:",
    "  ✅ Convergence 10x lebih cepat (gen 20 vs 200)",
    "  ✅ Final fitness 155 poin lebih tinggi",
    "  ✅ Selection pressure optimal untuk constraint ketat",
    "",
    "Roulette Selection:",
    "  ⚠️ Terlalu 'demokratis' untuk skenario extreme",
    "  ⚠️ Solusi suboptimal masih dapat breeding chance",
    "  ⚠️ Konvergensi lambat dengan 100 SKU",
]

for i, insight in enumerate(insights, start=13):
    ws3.cell(row=i, column=1, value=insight)
    ws3.merge_cells(f'A{i}:F{i}')
    if '✅' in insight:
        ws3.cell(row=i, column=1).font = Font(color="008000")
    elif '⚠️' in insight:
        ws3.cell(row=i, column=1).font = Font(color="FF6600")

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 15

# ============================================================================
# SHEET 4: PENJELASAN METODE
# ============================================================================

ws4 = wb.create_sheet("4. Penjelasan Metode")

ws4['A1'] = "PENJELASAN METODE SELEKSI"
ws4['A1'].font = Font(bold=True, size=14, color="004230")
ws4.merge_cells('A1:D1')

# Roulette explanation
ws4['A3'] = "1. ROULETTE WHEEL SELECTION"
ws4['A3'].font = Font(bold=True, size=12, color="0D8564")
ws4.merge_cells('A3:D3')

roulette_explanation = [
    "Konsep: Seperti roda roulette, area setiap individu sebanding dengan fitness",
    "",
    "Cara Kerja:",
    "  1. Hitung total fitness populasi",
    "  2. Probabilitas = fitness_individu / total_fitness",
    "  3. Spin roda (random 0-1), pilih individu sesuai area yang jatuh",
    "",
    "Karakteristik:",
    "  ✅ Proporsional: fitness tinggi = chance besar",
    "  ✅ Demokratis: semua punya kesempatan",
    "  ❌ Lambat konvergen: individu jelek masih breeding",
    "  ❌ Sensitive ke fitness scaling",
]

for i, text in enumerate(roulette_explanation, start=4):
    ws4.cell(row=i, column=1, value=text)
    ws4.merge_cells(f'A{i}:D{i}')

# Tournament explanation
ws4['A18'] = "2. TOURNAMENT SELECTION"
ws4['A18'].font = Font(bold=True, size=12, color="0D8564")
ws4.merge_cells('A18:D18')

tournament_explanation = [
    "Konsep: Seperti pertandingan olahraga - pilih random, yang terbaik menang",
    "",
    "Cara Kerja (Tournament Size = 3):",
    "  1. Pilih random 3 individu dari populasi",
    "  2. Bandingkan fitness ketiganya",
    "  3. Yang fitness tertinggi menjadi parent",
    "  4. Ulangi untuk parent berikutnya",
    "",
    "Karakteristik:",
    "  ✅ Selection pressure tinggi: fokus ke yang terbaik",
    "  ✅ Cepat konvergen: efisien untuk constraint ketat",
    "  ✅ Tidak perlu hitung total fitness",
    "  ✅ Tunable via tournament size",
    "  ❌ Kurang diverse: individu jelek jarang breeding",
]

for i, text in enumerate(tournament_explanation, start=19):
    ws4.cell(row=i, column=1, value=text)
    ws4.merge_cells(f'A{i}:D{i}')

ws4.column_dimensions['A'].width = 70

# ============================================================================
# SHEET 5: FITNESS FUNCTION
# ============================================================================

ws5 = wb.create_sheet("5. Fitness Function")

ws5['A1'] = "FITNESS FUNCTION COMPONENTS"
ws5['A1'].font = Font(bold=True, size=14, color="004230")
ws5.merge_cells('A1:E1')

ws5['A3'] = "Total Fitness = FC_CAP + FC_CAT + FC_AFF + FC_SPLIT"
ws5['A3'].font = Font(italic=True, size=11)
ws5.merge_cells('A3:E3')

# Components table
components = [
    ['Komponen', 'Bobot', 'Kondisi Terpenuhi', 'Kondisi Tidak Terpenuhi', 'Prioritas'],
    ['FC_CAP (Capacity)', '40 poin', 'Kapasitas cell mencukupi', 'Kapasitas tidak cukup (0 poin)', '1 (Highest)'],
    ['FC_CAT (Category)', '30 poin', 'SKU sesuai zona kategori', 'Zona tidak sesuai (0 poin)\nZona Mixed (15 poin)', '2'],
    ['FC_AFF (Affinity)', '20 poin', 'Ada barang sejenis di cell adjacent', 'Tidak ada barang sejenis (0 poin)\nSame row (10 poin)', '3'],
    ['FC_SPLIT (Split)', '10 poin', 'SKU hanya di satu lokasi', 'SKU split di beberapa lokasi (0 poin)', '4 (Lowest)'],
]

for i, row in enumerate(components, start=5):
    for j, value in enumerate(row, start=1):
        cell = ws5.cell(row=i, column=j, value=value)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if i == 5:
            cell.fill = subheader_fill
            cell.font = subheader_font

# Example calculation
ws5['A11'] = "📊 CONTOH PERHITUNGAN"
ws5['A11'].font = Font(bold=True, size=12, color="0D8564")

example = [
    "SKU: PAINT-001 (quantity: 8 units, category: Paint)",
    "Cell: 2B (capacity: 10/35 remaining, zone: Paint)",
    "",
    "Evaluasi:",
    "  FC_CAP = 40  (8 ≤ 10, kapasitas cukup) ✅",
    "  FC_CAT = 30  (Paint == Paint, kategori sesuai) ✅",
    "  FC_AFF = 20  (Cell 1B & 3B juga Paint, adjacent) ✅",
    "  FC_SPLIT = 10  (PAINT-001 hanya di cell 2B) ✅",
    "",
    "Total Fitness Gene = 40 + 30 + 20 + 10 = 100 (OPTIMAL!)",
]

for i, text in enumerate(example, start=12):
    ws5.cell(row=i, column=1, value=text)
    ws5.merge_cells(f'A{i}:E{i}')
    if '✅' in text:
        ws5.cell(row=i, column=1).font = Font(color="008000")

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 12
ws5.column_dimensions['C'].width = 25
ws5.column_dimensions['D'].width = 25
ws5.column_dimensions['E'].width = 15
ws5.row_dimensions[6].height = 30
ws5.row_dimensions[7].height = 40
ws5.row_dimensions[8].height = 40
ws5.row_dimensions[9].height = 30

# ============================================================================
# SAVE WORKBOOK
# ============================================================================

filename = 'GA_Warehouse_Optimization_Results.xlsx'
wb.save(filename)

print(f"\n✅ Excel documentation created successfully!")
print(f"📁 File: {filename}")
print(f"\n📊 Contains 5 sheets:")
print(f"   1. Executive Summary - Rekomendasi & Performance Summary")
print(f"   2. Results 25 SKUs - Hasil detail skenario 25 SKU")
print(f"   3. Results 100 SKUs - Head-to-head comparison extreme")
print(f"   4. Penjelasan Metode - Roulette vs Tournament")
print(f"   5. Fitness Function - Komponen & contoh perhitungan")
